#!/usr/bin/env python3
"""
策划案Excel生成脚本 v3 — 基于模板复制+样式克隆

核心思路：
  1. 复制 assets/template.xlsx 作为基底
  2. 清空所有Sheet中的示例数据
  3. 使用 theme 颜色精确还原模板视觉风格
  4. 根据JSON定义填充新内容
  5. 按需增删Sheet

Usage:
    python generate_design_doc.py <json_file_path> [output_dir]
"""

import json
import sys
import os
import copy
import shutil
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# 模板样式定义（精确复制自【规范】系统策划案模板.xlsx）
# ============================================================================

def _theme_color(theme, tint=0.0):
    c = Color(theme=theme, tint=tint)
    return c

# 文字颜色
CLR_WHITE = _theme_color(0)
CLR_BLACK = _theme_color(1)
CLR_BLUE_TEXT = _theme_color(4)
CLR_BLACK_RGB = Color(rgb='FF000000')

# 填充颜色
FILL_SHEET_TITLE = PatternFill(fill_type='solid')
FILL_SHEET_TITLE.fgColor = _theme_color(4, -0.5)

FILL_SECTION_TITLE = PatternFill(fill_type='solid')
FILL_SECTION_TITLE.fgColor = _theme_color(4, -0.25)

FILL_VERSION_VALUE = PatternFill(fill_type='solid')
FILL_VERSION_VALUE.fgColor = _theme_color(3, 0.9)

FILL_TABLE_HEADER = PatternFill(fill_type='solid')
FILL_TABLE_HEADER.fgColor = _theme_color(3, 0.5)

FILL_TABLE_HEADER_LIGHT = PatternFill(fill_type='solid')
FILL_TABLE_HEADER_LIGHT.fgColor = _theme_color(3, 0.6)

FILL_CONFIG_HEADER = PatternFill(fill_type='solid')
FILL_CONFIG_HEADER.fgColor = _theme_color(0, -0.15)

FILL_RED_DOT_HEADER = PatternFill(fill_type='solid')
FILL_RED_DOT_HEADER.fgColor = _theme_color(0, -0.05)

FILL_ALT_ROW = PatternFill(fill_type='solid')
FILL_ALT_ROW.fgColor = _theme_color(0, -0.05)

# 字体
FONT_SHEET_TITLE = Font(name='等线', size=22, bold=True, color=CLR_WHITE)
FONT_SECTION_TITLE = Font(name='等线', size=14, bold=True, color=CLR_WHITE)
FONT_VERSION_LABEL = Font(name='等线', size=12, bold=True, color=CLR_WHITE)
FONT_VERSION_VALUE = Font(name='等线', size=11, color=CLR_BLACK)
FONT_BODY = Font(name='等线', size=11, color=CLR_BLACK)
FONT_BODY_RGB = Font(name='等线', size=11, color=CLR_BLACK_RGB)
FONT_COMMENT = Font(name='等线', size=11, color=CLR_BLUE_TEXT)
FONT_TABLE_HEADER = Font(name='等线', size=11, bold=True, color=CLR_WHITE)
FONT_TABLE_HEADER_DARK = Font(name='等线', size=12, bold=True, color=CLR_BLACK)
FONT_CONFIG_HEADER = Font(name='等线', size=11, color=CLR_BLACK)
FONT_TRACKING_SUBTITLE = Font(name='等线', size=11, bold=True, color=CLR_WHITE)
FONT_SUBTITLE = Font(name='等线', size=11, bold=True, color=CLR_BLACK)
FONT_MARKER = Font(name='Segoe UI Symbol', size=11, color=CLR_BLACK)
FONT_SUMMARY = Font(name='等线', size=11, color=CLR_BLACK)
FONT_LINK = Font(name='等线', size=11, color=Color(rgb='FF4472C4'), underline='single')
FONT_IMG_PLACEHOLDER = Font(name='等线', size=11, color=Color(rgb='FF4472C4'), italic=True)

# 对齐
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center')
ALIGN_LEFT_CENTER_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_LEFT_TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_CENTER_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_CENTER_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT_CENTER = Alignment(horizontal='right', vertical='center')

# 边框
THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER = Border()

# 行高常量
ROW_H_SHEET_TITLE = 34.5
ROW_H_SECTION_TITLE = 18.0
ROW_H_VERSION_LABEL = 15.75
ROW_H_VERSION_VALUE = 14.25
ROW_H_BODY = 14.25
ROW_H_TABLE_HEADER = 15.75
ROW_H_TABLE_DATA = 20.1
ROW_H_TABLE_DATA_TALL = 27.75
ROW_H_IMG = 150


# ============================================================================
# 辅助函数
# ============================================================================

def write_cell(ws, row, col, value, font=None, alignment=None, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    return cell


def merge_write(ws, start_row, start_col, end_row, end_col, value,
                font=None, alignment=None, fill=None, border=None, row_height=None):
    if start_row != end_row or start_col != end_col:
        if border:
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    b_left = border.left if c == start_col else Side()
                    b_right = border.right if c == end_col else Side()
                    b_top = border.top if r == start_row else Side()
                    b_bottom = border.bottom if r == end_row else Side()
                    ws.cell(row=r, column=c).border = Border(
                        left=b_left, right=b_right, top=b_top, bottom=b_bottom
                    )
        if fill:
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).fill = fill
        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=end_row, end_column=end_col
        )
    cell = write_cell(ws, start_row, start_col, value, font, alignment, fill, border)
    if row_height:
        ws.row_dimensions[start_row].height = row_height
    return cell


def write_sheet_title(ws, row, text, end_col=21):
    merge_write(ws, row, 2, row, end_col, text,
                font=FONT_SHEET_TITLE, fill=FILL_SHEET_TITLE,
                alignment=ALIGN_LEFT_CENTER, row_height=ROW_H_SHEET_TITLE)
    return row + 1


def write_section_title(ws, row, text, end_col=12):
    merge_write(ws, row, 2, row, end_col, text,
                font=FONT_SECTION_TITLE, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_LEFT_CENTER, row_height=ROW_H_SECTION_TITLE)
    return row + 1


def write_comment(ws, row, text):
    write_cell(ws, row, 2, text, font=FONT_COMMENT, alignment=ALIGN_LEFT_CENTER)
    return row + 1


def write_body_text(ws, row, text, start_col=2, merge_cols=0):
    if merge_cols > 0:
        merge_write(ws, row, start_col, row, start_col + merge_cols - 1, text,
                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP)
    else:
        write_cell(ws, row, start_col, text, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_BODY
    return row + 1


def write_tracking_subtitle(ws, row, text, end_col=5):
    merge_write(ws, row, 2, row, end_col, text,
                font=FONT_TRACKING_SUBTITLE, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
    return row + 1


# ============================================================================
# 规则写入（▸/--层级系统）
# ============================================================================

def write_rules(ws, start_row, rules, base_col=2):
    row = start_row
    for rule in rules:
        level = rule.get('level', 1)
        text = rule.get('text', '')
        marker = rule.get('marker', '▸' if level == 1 else '--')

        marker_col = base_col + (level - 1)
        text_col = marker_col + 1

        if marker in ('▸', '▶'):
            write_cell(ws, row, marker_col, marker, font=FONT_MARKER, alignment=ALIGN_RIGHT_CENTER)
        else:
            write_cell(ws, row, marker_col, marker, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)

        write_cell(ws, row, text_col, text, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
        ws.row_dimensions[row].height = ROW_H_BODY
        row += 1

        children = rule.get('children', [])
        if children:
            row = write_rules(ws, row, children, base_col=base_col)

    return row


# ============================================================================
# 表格写入
# ============================================================================

def write_data_table(ws, start_row, start_col, headers, rows,
                     header_font=FONT_TABLE_HEADER,
                     header_fill=FILL_TABLE_HEADER,
                     header_align=ALIGN_LEFT_CENTER,
                     data_font=FONT_BODY,
                     data_align=ALIGN_LEFT_CENTER_WRAP,
                     col_merges=None,
                     header_col_merges=None,
                     row_height=ROW_H_TABLE_DATA):
    actual_merges = header_col_merges or col_merges

    # 写表头
    for i, header in enumerate(headers):
        phys_col = start_col + i
        if actual_merges and i in actual_merges:
            sc, ec = actual_merges[i]
            merge_write(ws, start_row, sc, start_row, ec, header,
                        font=header_font, fill=header_fill,
                        alignment=header_align, border=THIN_BORDER)
        else:
            if actual_merges:
                skip = False
                for mi, (sc, ec) in actual_merges.items():
                    if mi != i and sc <= phys_col <= ec:
                        skip = True
                        break
                if skip:
                    continue
            write_cell(ws, start_row, phys_col, header,
                       font=header_font, fill=header_fill,
                       alignment=header_align, border=THIN_BORDER)
    ws.row_dimensions[start_row].height = ROW_H_TABLE_HEADER

    # 写数据行
    data_merges = col_merges or {}
    for r_idx, row_data in enumerate(rows):
        cur_row = start_row + 1 + r_idx
        for c_idx, value in enumerate(row_data):
            phys_col = start_col + c_idx
            if data_merges and c_idx in data_merges:
                sc, ec = data_merges[c_idx]
                merge_write(ws, cur_row, sc, cur_row, ec, value,
                            font=data_font, alignment=data_align, border=THIN_BORDER)
            else:
                if data_merges:
                    skip = False
                    for mi, (sc, ec) in data_merges.items():
                        if mi != c_idx and sc <= phys_col <= ec:
                            skip = True
                            break
                    if skip:
                        continue
                write_cell(ws, cur_row, phys_col, value,
                           font=data_font, alignment=data_align, border=THIN_BORDER)
        ws.row_dimensions[cur_row].height = row_height

    return start_row + 1 + len(rows)


# ============================================================================
# 各Sheet生成函数
# ============================================================================

def clear_sheet(ws):
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.border = Border()


def gen_version_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 30.0
    ws.column_dimensions['A'].width = 9.0

    row = 2
    merge_write(ws, row, 2, row, 4, '方案版本号',
                font=FONT_VERSION_LABEL, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_LABEL
    row += 1

    merge_write(ws, row, 2, row, 4, data.get('version', '1.0'),
                font=FONT_VERSION_VALUE, fill=FILL_VERSION_VALUE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    merge_write(ws, row, 2, row, 4, '设计人',
                font=FONT_VERSION_LABEL, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_LABEL
    row += 1

    merge_write(ws, row, 2, row, 4, data.get('author', 'Apakoh(喻骋远)'),
                font=FONT_VERSION_VALUE, fill=FILL_VERSION_VALUE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    merge_write(ws, row, 2, row, 4, '方案内容',
                font=FONT_VERSION_LABEL, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_LABEL
    row += 1

    merge_write(ws, row, 2, row, 4, data.get('title', ''),
                font=FONT_VERSION_VALUE, fill=FILL_VERSION_VALUE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    ws.row_dimensions[row].height = ROW_H_VERSION_VALUE
    row += 1

    merge_write(ws, row, 2, row, 11, '概述',
                font=FONT_VERSION_LABEL, fill=FILL_SECTION_TITLE,
                alignment=ALIGN_LEFT_CENTER)
    ws.row_dimensions[row].height = ROW_H_VERSION_LABEL
    row += 1

    summary = data.get('summary', '')
    end_row = row + 33
    merge_write(ws, row, 2, end_row, 11, summary,
                font=FONT_SUMMARY, alignment=ALIGN_LEFT_TOP_WRAP)
    for r in range(row, end_row + 1):
        ws.row_dimensions[r].height = ROW_H_VERSION_VALUE


def gen_interaction_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 14.25

    row = 2
    row = write_sheet_title(ws, row, '交互说明', end_col=21)
    row += 1

    # 新增美术资产量评估
    art_assets = data.get('art_assets', [])
    row = write_section_title(ws, row, '新增美术资产量评估', end_col=12)

    comments = data.get('art_asset_comments', [])
    for c in comments:
        row = write_comment(ws, row, c)

    row += 1

    if art_assets:
        write_cell(ws, row, 2, '分类', font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER_LIGHT,
                   alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 3, row, 5, '详细',
                    font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER_LIGHT,
                    alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        write_cell(ws, row, 6, '数量', font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER_LIGHT,
                   alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 7, row, 23, '说明',
                    font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER_LIGHT,
                    alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_DATA
        row += 1

        for asset in art_assets:
            write_cell(ws, row, 2, asset.get('category', ''),
                       font=FONT_BODY_RGB, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
            merge_write(ws, row, 3, row, 5, asset.get('detail', ''),
                        font=FONT_BODY_RGB, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
            write_cell(ws, row, 6, asset.get('count', ''),
                       font=FONT_BODY_RGB, alignment=ALIGN_CENTER_CENTER, border=THIN_BORDER)
            merge_write(ws, row, 7, row, 23, asset.get('note', ''),
                        font=FONT_BODY_RGB, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_TABLE_DATA
            row += 1

    row += 1

    # 主要界面流转
    flow = data.get('flow_overview', {})
    row = write_section_title(ws, row, '主要界面流转（概览）', end_col=12)

    flow_comments = data.get('flow_comments', [])
    for c in flow_comments:
        row = write_comment(ws, row, c)

    if flow:
        for link in flow.get('links', []):
            cell = write_cell(ws, row, 2, link, font=FONT_LINK, alignment=ALIGN_LEFT_CENTER)
            cell.hyperlink = link
            row += 1
        desc = flow.get('description', '')
        if desc:
            row = write_body_text(ws, row, desc, start_col=2, merge_cols=10)
        merge_write(ws, row, 2, row, 12, '[图片占位: 主要界面流转图]',
                    font=FONT_IMG_PLACEHOLDER, alignment=ALIGN_CENTER_CENTER,
                    row_height=ROW_H_IMG)
        row += 2

    row += 1

    # 主要界面详细说明
    detail_views = data.get('detail_views', [])
    row = write_section_title(ws, row, '主要界面详细说明', end_col=12)

    detail_comments = data.get('detail_comments', [])
    for c in detail_comments:
        row = write_comment(ws, row, c)

    for view in detail_views:
        view_name = view.get('view_name', '')
        row += 1
        write_cell(ws, row, 2, view_name, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
        row += 1
        for link in view.get('links', []):
            cell = write_cell(ws, row, 2, link, font=FONT_LINK, alignment=ALIGN_LEFT_CENTER)
            cell.hyperlink = link
            row += 1
        desc = view.get('description', '')
        if desc:
            row = write_body_text(ws, row, desc, start_col=2, merge_cols=10)
        merge_write(ws, row, 2, row, 12, f'[图片占位: {view_name}UE注解图]',
                    font=FONT_IMG_PLACEHOLDER, alignment=ALIGN_CENTER_CENTER,
                    row_height=ROW_H_IMG)
        row += 2


def gen_feature_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 14.25

    row = 2
    row = write_sheet_title(ws, row, '功能说明', end_col=21)
    row += 1

    for section in data.get('sections', []):
        row = write_section_title(ws, row, section.get('title', ''), end_col=11)

        for subsection in section.get('subsections', [{}]):
            sub_title = subsection.get('title', '')
            if sub_title:
                write_cell(ws, row, 2, sub_title, font=FONT_SUBTITLE, alignment=ALIGN_LEFT_CENTER)
                row += 1

            rules = subsection.get('rules', [])
            if rules:
                row = write_rules(ws, row, rules, base_col=2)

            paragraphs = subsection.get('paragraphs', [])
            for p in paragraphs:
                write_cell(ws, row, 3, p, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
                row += 1

            tables = subsection.get('tables', [])
            for tbl in tables:
                headers = tbl.get('headers', [])
                tbl_rows = tbl.get('rows', [])
                if headers:
                    row = write_data_table(
                        ws, row, 3, headers, tbl_rows,
                        header_font=FONT_BODY,
                        header_fill=FILL_CONFIG_HEADER,
                        header_align=ALIGN_LEFT_CENTER_WRAP,
                        data_font=FONT_BODY,
                        data_align=ALIGN_LEFT_CENTER_WRAP,
                        row_height=16.5
                    )
                row += 1

            notes = subsection.get('notes', [])
            for note in notes:
                write_cell(ws, row, 3, note, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
                row += 1

            images = subsection.get('images', [])
            for img in images:
                merge_write(ws, row, 2, row, 12, f'[图片占位: {img}]',
                            font=FONT_IMG_PLACEHOLDER, alignment=ALIGN_CENTER_CENTER,
                            row_height=ROW_H_IMG)
                row += 2

            row += 2  # 段落间距

        row += 1


def gen_red_dot_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 20.1

    row = 2
    row = write_sheet_title(ws, row, '红点透传逻辑', end_col=15)
    row += 1

    entries = data.get('entries', [])

    header_labels = ['红点情况描述', '描述出现红点的UI控件名称/路径', '红点触发条件',
                     '红点消除条件', '红点类型/数字标等说明']
    header_cols = [(2, 4), (5, 6), (7, 8), (9, 10), (11, 12)]

    for label, (sc, ec) in zip(header_labels, header_cols):
        merge_write(ws, row, sc, row, ec, label,
                    font=FONT_TABLE_HEADER_DARK, fill=FILL_RED_DOT_HEADER,
                    alignment=ALIGN_CENTER_CENTER, border=THIN_BORDER)
    ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
    row += 1

    flow_desc = data.get('flow_description', '')
    if flow_desc:
        merge_write(ws, row, 2, row, 12, flow_desc,
                    font=FONT_BODY, alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        ws.row_dimensions[row].height = 158.25
        row += 1

    for entry in entries:
        desc = entry.get('description', '')
        ui_path = entry.get('ui_path', '')
        trigger = entry.get('trigger', '')
        dismiss = entry.get('dismiss', '')
        dot_type = entry.get('dot_type', '★')

        merge_write(ws, row, 2, row, 4, desc,
                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 5, row, 6, ui_path,
                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 7, row, 8, trigger,
                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 9, row, 10, dismiss,
                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
        merge_write(ws, row, 11, row, 12, dot_type,
                    font=Font(name='等线', size=22, color=CLR_BLACK),
                    alignment=ALIGN_CENTER_CENTER, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_DATA_TALL
        row += 1


def gen_config_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 20.1

    row = 2
    row = write_sheet_title(ws, row, '配置需求', end_col=15)
    row += 1

    for table_def in data.get('tables', []):
        table_name = table_def.get('table_name', '')
        is_new = table_def.get('is_new', True)
        title_text = table_name
        if not is_new:
            title_text = f'{table_name}（字段拓展）'

        row = write_section_title(ws, row, title_text, end_col=12)

        fields = table_def.get('fields', [])
        if fields:
            # 检查是否有 table_type 字段
            has_table_type = any(f.get('table_type') for f in fields)

            if has_table_type:
                headers = ['字段名', 'EN', '数据类型', '打表类型', '字段说明']
                header_merges = {0: (2, 3), 4: (7, 13)}

                for i, header in enumerate(headers):
                    if i in header_merges:
                        sc, ec = header_merges[i]
                        merge_write(ws, row, sc, row, ec, header,
                                    font=FONT_CONFIG_HEADER, fill=FILL_CONFIG_HEADER,
                                    alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    else:
                        col_map = {1: 4, 2: 5, 3: 6}
                        phys_col = col_map.get(i, 2 + i)
                        write_cell(ws, row, phys_col, header,
                                   font=FONT_CONFIG_HEADER, fill=FILL_CONFIG_HEADER,
                                   alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                ws.row_dimensions[row].height = ROW_H_TABLE_DATA
                row += 1

                for field in fields:
                    merge_write(ws, row, 2, row, 3, field.get('field_name', ''),
                                font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    write_cell(ws, row, 4, field.get('en_name', ''),
                               font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    write_cell(ws, row, 5, field.get('data_type', ''),
                               font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    write_cell(ws, row, 6, field.get('table_type', ''),
                               font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    merge_write(ws, row, 7, row, 13, field.get('description', ''),
                                font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
                    ws.row_dimensions[row].height = ROW_H_TABLE_DATA
                    row += 1
            else:
                # 无打表类型的格式（拓展字段等）
                headers = ['字段名（新增）', 'EN', '数据类型', '字段说明']
                header_merges = {0: (2, 3), 3: (6, 13)}

                for i, header in enumerate(headers):
                    if i in header_merges:
                        sc, ec = header_merges[i]
                        merge_write(ws, row, sc, row, ec, header,
                                    font=FONT_CONFIG_HEADER, fill=FILL_CONFIG_HEADER,
                                    alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    else:
                        col_map = {1: 4, 2: 5}
                        phys_col = col_map.get(i, 2 + i)
                        write_cell(ws, row, phys_col, header,
                                   font=FONT_CONFIG_HEADER, fill=FILL_CONFIG_HEADER,
                                   alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                ws.row_dimensions[row].height = ROW_H_TABLE_DATA
                row += 1

                for field in fields:
                    merge_write(ws, row, 2, row, 3, field.get('field_name', ''),
                                font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    write_cell(ws, row, 4, field.get('en_name', ''),
                               font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    write_cell(ws, row, 5, field.get('data_type', ''),
                               font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                    merge_write(ws, row, 6, row, 13, field.get('description', ''),
                                font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
                    ws.row_dimensions[row].height = ROW_H_TABLE_DATA
                    row += 1

            # 附加说明文本
            extra_notes = table_def.get('notes', [])
            for note in extra_notes:
                row += 1
                write_cell(ws, row, 2, note, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP)
                row += 1

        row += 2


def gen_ip_icon_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 20.1
    ws.column_dimensions['A'].width = 9.0

    row = 2
    row = write_sheet_title(ws, row, '包装需求', end_col=21)

    global_comment = data.get('global_comment', '')
    if global_comment:
        row = write_comment(ws, row, global_comment)
    row += 1

    # 道具需求
    item_reqs = data.get('item_requirements', [])
    if item_reqs:
        row = write_section_title(ws, row, '道具需求', end_col=12)

        h_labels = ['ItemID', '道具名（IP填）\n≤5字', '道具类型', '道具icon（美术制作）',
                    '道具icon示意参考', '道具说明（IP填）\n≤30字']
        h_merges = [(2, 2), (3, 4), (5, 7), (8, 20), (21, 24), (25, 29)]
        for label, (sc, ec) in zip(h_labels, h_merges):
            merge_write(ws, row, sc, row, ec, label,
                        font=Font(name='等线', size=11, bold=True, color=CLR_WHITE),
                        fill=FILL_TABLE_HEADER_LIGHT,
                        alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        ws.row_dimensions[row].height = 38.25
        row += 1

        for item in item_reqs:
            vals = [
                item.get('item_id', ''),
                item.get('item_name', ''),
                item.get('item_type', ''),
                item.get('icon_desc', ''),
                item.get('icon_ref', ''),
                item.get('item_desc', ''),
            ]
            for v, (sc, ec) in zip(vals, h_merges):
                merge_write(ws, row, sc, row, ec, v,
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = 135.0
            row += 1

        row += 2

    # IP包装文案需求
    ip_reqs = data.get('ip_requirements', [])
    if ip_reqs:
        row = write_section_title(ws, row, 'IP包装文案需求', end_col=12)

        h_labels = ['功能/包装位置', '包装类型', '数量', '文案包装（IP填写）', '字限建议']
        h_merges_ip = [(2, 3), (4, 5), (6, 7), (8, 13), (15, 17)]
        for label, (sc, ec) in zip(h_labels, h_merges_ip):
            merge_write(ws, row, sc, row, ec, label,
                        font=Font(name='等线', size=11, bold=True, color=CLR_WHITE),
                        fill=FILL_TABLE_HEADER_LIGHT,
                        alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
        row += 1

        for req in ip_reqs:
            vals = [
                req.get('content', ''),
                req.get('pack_type', ''),
                req.get('count', ''),
                req.get('text_pack', ''),
                req.get('char_limit', ''),
            ]
            for v, (sc, ec) in zip(vals, h_merges_ip):
                merge_write(ws, row, sc, row, ec, v,
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_TABLE_DATA
            row += 1

        row += 2

    # 功能名包装需求
    naming_reqs = data.get('naming_requirements', [])
    if naming_reqs:
        row = write_section_title(ws, row, '功能名包装需求', end_col=12)

        h_labels2 = ['包装内容', '包装类型', '数量', '文案包装（IP填写）']
        h_merges2 = [(2, 5), (6, 8), (9, 12), (13, 15)]
        for i, label in enumerate(h_labels2):
            if i < len(h_merges2):
                sc, ec = h_merges2[i]
                merge_write(ws, row, sc, row, ec, label,
                            font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER,
                            alignment=ALIGN_CENTER_CENTER_WRAP, border=THIN_BORDER)
        ws.row_dimensions[row].height = 40.5
        row += 1

        for req in naming_reqs:
            vals = [
                req.get('content', ''),
                req.get('pack_type', ''),
                req.get('count', ''),
                req.get('text_pack', ''),
            ]
            for i, v in enumerate(vals):
                if i < len(h_merges2):
                    sc, ec = h_merges2[i]
                    fill = FILL_ALT_ROW if i == 3 else None
                    merge_write(ws, row, sc, row, ec, v,
                                font=FONT_BODY, alignment=ALIGN_CENTER_CENTER_WRAP,
                                border=THIN_BORDER, fill=fill)
            ws.row_dimensions[row].height = ROW_H_TABLE_DATA
            row += 1

        row += 2

    # 关卡文案需求
    level_names = data.get('level_names', [])
    if level_names:
        row = write_section_title(ws, row, '关卡文案需求', end_col=12)

        h_labels_l = ['关卡序号', '关卡名（IP填）\n建议≤5字']
        h_merges_l = [(2, 5), (6, 10)]

        # 如果有分组（如普通/困难），按两列排列
        groups = {}
        for ln in level_names:
            g = ln.get('group', '默认')
            if g not in groups:
                groups[g] = []
            groups[g].append(ln)

        if len(groups) > 1:
            # 多组横向排列
            group_list = list(groups.items())
            cols_per_group = [(2, 5, 6, 10), (11, 14, 15, 19)]
            for gi, (gname, items) in enumerate(group_list):
                if gi >= len(cols_per_group):
                    break
                c1s, c1e, c2s, c2e = cols_per_group[gi]
                merge_write(ws, row, c1s, row, c1e, '关卡序号',
                            font=Font(name='等线', size=11, bold=True, color=CLR_BLACK),
                            fill=FILL_CONFIG_HEADER,
                            alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                merge_write(ws, row, c2s, row, c2e, f'关卡名（IP填）\n建议≤5字',
                            font=Font(name='等线', size=11, bold=True, color=CLR_BLACK),
                            fill=FILL_CONFIG_HEADER,
                            alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
            row += 1

            max_items = max(len(items) for _, items in group_list)
            for i in range(max_items):
                for gi, (gname, items) in enumerate(group_list):
                    if gi >= len(cols_per_group):
                        break
                    c1s, c1e, c2s, c2e = cols_per_group[gi]
                    if i < len(items):
                        merge_write(ws, row, c1s, row, c1e, items[i].get('label', ''),
                                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                        merge_write(ws, row, c2s, row, c2e, items[i].get('name', ''),
                                    font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                ws.row_dimensions[row].height = ROW_H_BODY
                row += 1
        else:
            for label, (sc, ec) in zip(h_labels_l, h_merges_l):
                merge_write(ws, row, sc, row, ec, label,
                            font=Font(name='等线', size=11, bold=True, color=CLR_BLACK),
                            fill=FILL_CONFIG_HEADER,
                            alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
            row += 1

            for ln in level_names:
                merge_write(ws, row, 2, row, 5, ln.get('label', ''),
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                merge_write(ws, row, 6, row, 10, ln.get('name', ''),
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
                ws.row_dimensions[row].height = ROW_H_BODY
                row += 1

        row += 2

    # 对话文案规划
    dialogues = data.get('dialogues', [])
    if dialogues:
        row = write_section_title(ws, row, '对话文案规划', end_col=12)

        h_labels_d = ['文案包装（IP填写）', '语境/建议触发条件']
        h_merges_d = [(2, 8), (9, 15)]
        for label, (sc, ec) in zip(h_labels_d, h_merges_d):
            merge_write(ws, row, sc, row, ec, label,
                        font=Font(name='等线', size=12, bold=True, color=CLR_WHITE),
                        fill=PatternFill(fill_type='solid', fgColor=_theme_color(3, 0.5)),
                        alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
        row += 1

        for d in dialogues:
            merge_write(ws, row, 2, row, 8, d.get('text', ''),
                        font=FONT_BODY, alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
            merge_write(ws, row, 9, row, 15, d.get('condition', ''),
                        font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_BODY
            row += 1

        row += 2

    # 玩法/功能说明文案
    func_texts = data.get('function_texts', [])
    if func_texts:
        row = write_section_title(ws, row, '玩法/功能说明文案', end_col=12)

        for ft in func_texts:
            title = ft.get('title', '')
            content = ft.get('content', '')
            merge_write(ws, row, 2, row, 19, title,
                        font=Font(name='等线', size=11, bold=True), alignment=ALIGN_LEFT_CENTER_WRAP,
                        border=THIN_BORDER)
            row += 1
            end_row = row + max(0, len(content.split('\n')) - 1) + 5
            merge_write(ws, row, 2, end_row, 19, content,
                        font=FONT_BODY_RGB, alignment=ALIGN_LEFT_CENTER_WRAP,
                        border=THIN_BORDER)
            row = end_row + 2


def gen_tracking_gm_sheet(ws, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 14.25
    ws.column_dimensions['A'].width = 9.0

    row = 2
    row = write_sheet_title(ws, row, '埋点需求', end_col=22)
    row += 1

    # 新增快照
    snapshots = data.get('snapshots', [])
    if snapshots:
        row = write_tracking_subtitle(ws, row, '新增快照', end_col=5)

        snap_headers = ['Column（字段名）', '发送快照时机', '描述', '格式', '归类', '备注']
        snap_cols = [(2, 3), (4, 5), (6, 11), (12, 14), (15, 15), (16, 20)]
        for label, (sc, ec) in zip(snap_headers, snap_cols):
            merge_write(ws, row, sc, row, ec, label,
                        font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER,
                        alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
        row += 1

        for snap in snapshots:
            vals = [
                snap.get('field_name', ''),
                snap.get('timing', ''),
                snap.get('description', ''),
                snap.get('format', ''),
                snap.get('category', ''),
                snap.get('note', ''),
            ]
            for v, (sc, ec) in zip(vals, snap_cols):
                merge_write(ws, row, sc, row, ec, v,
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = ROW_H_TABLE_DATA
            row += 1

        row += 1

    # 新增埋点
    events = data.get('tracking_events', [])
    if events:
        row = write_tracking_subtitle(ws, row, '新增埋点', end_col=5)

        evt_headers = ['大类', '子类', '埋点归属\nclient/server', 'Event', '名称', '上报字段']
        evt_cols = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11), (12, 17)]
        for label, (sc, ec) in zip(evt_headers, evt_cols):
            merge_write(ws, row, sc, row, ec, label,
                        font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER,
                        alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
        row += 1

        prev_cat = None
        cat_start_row = row
        for i, evt in enumerate(events):
            cat = evt.get('category', '')
            vals = [
                cat,
                evt.get('sub_category', ''),
                evt.get('ownership', ''),
                evt.get('event', ''),
                evt.get('name', ''),
                evt.get('fields', ''),
            ]
            for j, (v, (sc, ec)) in enumerate(zip(vals, evt_cols)):
                if j == 0:
                    if cat != prev_cat:
                        if prev_cat is not None and i > 0 and cat_start_row < row:
                            ws.merge_cells(start_row=cat_start_row, start_column=2,
                                           end_row=row - 1, end_column=3)
                        cat_start_row = row
                        merge_write(ws, row, sc, row, ec, v,
                                    font=FONT_BODY, alignment=ALIGN_CENTER_CENTER_WRAP,
                                    border=THIN_BORDER)
                        prev_cat = cat
                    else:
                        merge_write(ws, row, sc, row, ec, '',
                                    font=FONT_BODY, alignment=ALIGN_CENTER_CENTER_WRAP,
                                    border=THIN_BORDER)
                else:
                    merge_write(ws, row, sc, row, ec, v,
                                font=FONT_BODY, alignment=ALIGN_CENTER_CENTER_WRAP,
                                border=THIN_BORDER)
            ws.row_dimensions[row].height = 48.0
            row += 1

        if prev_cat is not None and cat_start_row < row:
            if cat_start_row < row - 1:
                ws.merge_cells(start_row=cat_start_row, start_column=2,
                               end_row=row - 1, end_column=3)

    row += 2

    # GM需求
    row = write_sheet_title(ws, row, 'GM需求', end_col=22)
    row += 1

    gm_cmds = data.get('gm_commands', [])
    if gm_cmds:
        row = write_tracking_subtitle(ws, row, '新增GM指令需求', end_col=5)

        gm_headers = ['指令', '参数', '描述']
        gm_cols = [(2, 3), (4, 9), (10, 15)]
        for label, (sc, ec) in zip(gm_headers, gm_cols):
            merge_write(ws, row, sc, row, ec, label,
                        font=FONT_TABLE_HEADER, fill=FILL_TABLE_HEADER,
                        alignment=ALIGN_LEFT_CENTER, border=THIN_BORDER)
        ws.row_dimensions[row].height = ROW_H_TABLE_HEADER
        row += 1

        for cmd in gm_cmds:
            vals = [
                cmd.get('command', ''),
                cmd.get('params', ''),
                cmd.get('description', ''),
            ]
            for v, (sc, ec) in zip(vals, gm_cols):
                merge_write(ws, row, sc, row, ec, v,
                            font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP, border=THIN_BORDER)
            ws.row_dimensions[row].height = 31.5
            row += 1


def gen_generic_sheet(ws, sheet_name, data):
    clear_sheet(ws)
    ws.sheet_format.defaultRowHeight = 14.25

    row = 2
    blocks = data.get('blocks', [])
    for block in blocks:
        btype = block.get('type', 'paragraph')

        if btype == 'sheet_title':
            row = write_sheet_title(ws, row, block.get('text', ''), end_col=21)
        elif btype == 'section_title':
            row = write_section_title(ws, row, block.get('text', ''), end_col=12)
        elif btype == 'title':
            row = write_section_title(ws, row, block.get('text', ''), end_col=12)
        elif btype == 'subtitle':
            write_cell(ws, row, 2, block.get('text', ''), font=FONT_SUBTITLE, alignment=ALIGN_LEFT_CENTER)
            row += 1
        elif btype == 'paragraph':
            text = block.get('text', '')
            merge_cols = block.get('merge_cols', 10)
            merge_write(ws, row, 2, row, 2 + merge_cols - 1, text,
                        font=FONT_BODY, alignment=ALIGN_LEFT_CENTER_WRAP)
            row += 1
        elif btype == 'comment':
            row = write_comment(ws, row, block.get('text', ''))
        elif btype == 'link':
            text = block.get('text', '')
            url = block.get('url', '')
            if text:
                write_cell(ws, row, 2, text, font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
            if url:
                col = 3 if text else 2
                cell = write_cell(ws, row, col, url, font=FONT_LINK, alignment=ALIGN_LEFT_CENTER)
                cell.hyperlink = url
            row += 1
        elif btype == 'rule':
            row = write_rules(ws, row, [block], base_col=2)
        elif btype == 'rules':
            row = write_rules(ws, row, block.get('items', []), base_col=2)
        elif btype == 'table':
            headers = block.get('headers', [])
            tbl_rows = block.get('rows', [])
            if headers:
                row = write_data_table(ws, row, 2, headers, tbl_rows)
        elif btype == 'image_placeholder':
            desc = block.get('description', '图片')
            merge_write(ws, row, 2, row, 12, f'[图片占位: {desc}]',
                        font=FONT_IMG_PLACEHOLDER, alignment=ALIGN_CENTER_CENTER,
                        row_height=ROW_H_IMG)
            row += 1
        elif btype == 'empty':
            row += block.get('count', 1)
        else:
            write_cell(ws, row, 2, str(block.get('text', '')),
                       font=FONT_BODY, alignment=ALIGN_LEFT_CENTER)
            row += 1


# ============================================================================
# 主生成函数
# ============================================================================

SHEET_TYPE_MAP = {
    'version_info': ('版本号', gen_version_sheet),
    'interaction': ('交互', gen_interaction_sheet),
    'feature_spec': ('功能说明', gen_feature_sheet),
    'red_dot': ('红点逻辑', gen_red_dot_sheet),
    'config_spec': ('配置需求', gen_config_sheet),
    'ip_icon': ('IP&Icon需求', gen_ip_icon_sheet),
    'tracking_gm': ('埋点&GM需求', gen_tracking_gm_sheet),
}

TEMPLATE_SHEET_NAMES = ['版本号', '交互', '功能说明', '红点逻辑', '配置需求', 'IP&Icon需求', '埋点&GM需求']


def generate_design_doc(json_data, output_path, template_path=None):
    if 'file_name' not in json_data:
        raise ValueError("JSON定义缺少 'file_name' 字段")
    if 'sheets' not in json_data or not json_data['sheets']:
        raise ValueError("JSON定义缺少 'sheets' 字段或sheets为空")

    if template_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, '..', 'assets', 'template.xlsx')

    if os.path.exists(template_path):
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        print(f"Warning: 模板文件不存在({template_path})，将创建新工作簿")
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    needed_sheets = set()

    for sheet_def in json_data['sheets']:
        sheet_type = sheet_def.get('type', 'generic')
        sheet_data = sheet_def.get('data', {})
        custom_name = sheet_def.get('name', '')

        if sheet_type in SHEET_TYPE_MAP:
            default_name, generator = SHEET_TYPE_MAP[sheet_type]
            ws = None
            if default_name in wb.sheetnames:
                ws = wb[default_name]
            elif custom_name and custom_name in wb.sheetnames:
                ws = wb[custom_name]
            else:
                ws = wb.create_sheet(custom_name or default_name)

            generator(ws, sheet_data)
            needed_sheets.add(ws.title)
        else:
            actual_name = custom_name or f'Sheet_{len(needed_sheets)}'
            if actual_name in wb.sheetnames:
                ws = wb[actual_name]
            else:
                ws = wb.create_sheet(actual_name)
            gen_generic_sheet(ws, actual_name, sheet_data)
            needed_sheets.add(actual_name)

    for sname in list(wb.sheetnames):
        if sname not in needed_sheets:
            del wb[sname]

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"策划案Excel已生成: {os.path.abspath(output_path)}")
    return os.path.abspath(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_design_doc.py <json_file_path> [output_dir]")
        sys.exit(1)

    json_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '.'

    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
    except FileNotFoundError:
        print(f"Error: JSON文件不存在: {json_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: JSON解析失败: {e}")
        sys.exit(1)

    file_name = json_data.get('file_name', '策划案')
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'
    output_path = os.path.join(output_dir, file_name)

    try:
        result = generate_design_doc(json_data, output_path)
        print(f"生成成功: {result}")
    except Exception as e:
        print(f"Error: 生成失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
